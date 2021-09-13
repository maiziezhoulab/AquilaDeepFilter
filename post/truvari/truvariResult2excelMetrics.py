from openpyxl import Workbook, load_workbook
import os
import json
import argparse
from pathlib import Path


def list_(s):
    try:
        l = s.split(',')
        return l
    except:
        raise argparse.ArgumentTypeError("comma as separator")


parser = argparse.ArgumentParser(description='evaluate with Truvari')
parser.add_argument('--root_folder',  # /data/maiziezhou_lab/huyf/DeepSVFilter/results/10xweb_raw/0706ensemble/
                              type=str,
                              required=True,
                              help='path to root folder that contains all truvari metrics of a batch')
parser.add_argument('--output_excel',
                              type=str,
                              required=True,
                              help='path to output excel')
parser.add_argument('--ensemble',
                              type=bool,
                              default=False,
                              help='ensemble or batch')
parser.add_argument('--models',
                              type=list_,
                              default='densenet,efficientnet,mobilenet,resnet,xception',
                              help='all the candidate subfolders to use')

args = parser.parse_args()

if args.ensemble:  # only one sub folder if --ensemble is used
    ###############################
    # implement ...
    ###############################
    """headers"""
    wb = Workbook()
    ws = wb.active
    header_1 = ['TP','FP','FN','precision', 'recall', 'f1']
    ws.cell(row=1, column=1).value = 'threshold'
    # ws.cell(row=1, column=13).value = 'stlfr'
    for i in range(1,7):
        ws.cell(row=1, column=i+1, value=header_1[i-1])
        # ws.cell(row=1, column=i+13, value=header_1[i-1])
    wb.save(filename=args.output_excel)
    
    """content"""
    root_folder = args.root_folder
    alls = os.listdir(root_folder)
    dirs = []
    
    for f in alls:
        if Path(os.path.join(root_folder, f)).is_dir():
            dirs.append(os.path.join(root_folder, f))
    thres_dict = dict()
    for sub_dir in dirs:
        j_file = open(os.path.join(sub_dir, 'summary.txt'))
        data = json.loads(j_file.read())
        # print(data)
        prec = data['precision']
        recall = data['recall']
        f1 = data['f1']
        tp = data['TP-base']
        fp = data['FP']
        fn = data['FN']
        thres_ = float(sub_dir.split('/')[-1].split('_')[0][-3:])
        info_list = [tp, fp, fn, prec, recall, f1]
        thres_dict[float(thres_)] = info_list
    i = 2
    for e in thres_dict:
        workbook = load_workbook(filename=args.output_excel)
        sheet = workbook['Sheet']
        # print(e)
        # print(thres_dict[e])
        # exit(-1)
        sheet.cell(row=i, column=1, value = e)
        for j in range(len(thres_dict[e])):
            sheet.cell(row=i, column=2+j, value = thres_dict[e][j])
        # sheet.cell(row=i+2, column=13, value = keys[i])
        workbook.save(filename=args.output_excel)
        i += 1
    # print(args.models)
    # continue
else:
    root_folder = args.root_folder  # else than there are few subfolders that are densenet/ .../ .../ ...
    ################################
    #  modify the codes to relaxate the number of subfolders, 5 for normal, 3 for balanced data
    ################################
    models_ = args.models
    divide_folders = []
    # models = {0:'densenet',1:'efficientnet',2:'mobilenet',3:'resnet',4:'xception'}   # typo here, it is efficientnet not det
    models = dict()
    index_  = 0
    for m in models_:
        models[index_] = m
        divide_folders.append([])
        index_ += 1
    subfolders = [os.path.join(root_folder, o) for o in os.listdir(root_folder) if os.path.isdir(os.path.join(root_folder, o))]
    batch_ = subfolders[0][-4:]
    # print(subfolders)
    # print(batch_)
    # print(divide_folders)
    # print(subfolders)
    # print(models_)
    # print(len(subfolders), len(models_))
    if len(subfolders) != len(models_):
        print("double check the parameter --models. it does not match")
        exit(-1)
    else:
        print("all good")
    # exit(-1)
    for f in subfolders:
        # model_name = f
        p = os.path.join(root_folder, f)
        sub_subfolders = os.listdir(p)
        for subf in sub_subfolders:
            for i in range(len(divide_folders)):
                if subf.startswith(models[i]) and Path(os.path.join(p, subf)).is_dir():
                    divide_folders[i].append(os.path.join(p, subf))
                    break
                
        # if model_name.startswith(models[0]):
        #     divide_folders[0].append(f)
        # elif model_name.startswith(models[1]):
        #     divide_folders[1].append(f)
        # elif model_name.startswith(models[2]):
        #     divide_folders[2].append(f)
        # elif model_name.startswith(models[3]):
        #     divide_folders[3].append(f)
        # elif model_name.startswith(models[4]):
        #     divide_folders[4].append(f)
    
    # print(divide_folders)
    ############################
    # len(divide_folders[0]) = # of thresholds
    ##############################
    # print(len(divide_folders[0]))
    # exit(-1)
    # create_sheet('sid1')
    wb = Workbook()
    ws = wb.active
    header_1 = ['TP','FP','FN','precision', 'recall', 'f1', 'svtype', 'batch', 'tag', 'sample', 'lib', 'model']
    # ws.cell(row=1, column=1).value = '10xweb'
    # ws.cell(row=1, column=13).value = 'stlfr'
    for i in range(1,13):
        ws.cell(row=1, column=i+1, value=header_1[i-1])
    for n in range(len(divide_folders)-1):
        ws = wb.create_sheet('Sheet' + str(n+1))
    # for s in wb.sheetnames:
        for i in range(1,13):
            ws.cell(row=1, column=i+1, value=header_1[i-1])
            # ws.cell(row=1, column=i+13, value=header_1[i-1])
        # sheet = Workbook['Sheet']
        # print(wb.sheetnames)
        wb.save(filename=args.output_excel)
        # exit(-1)
    sheets_list = wb.sheetnames
    # print(sheets_list)
    # exit(-1)
    k = 0
    for temp_folders in divide_folders:
        model_n = models[k]
        print(model_n)
        
        
        
        thres_dict = dict()
        # t2_dict = dict()
        
        
        for f in temp_folders:
            print(f.split('/'))
            temp_split = f.split('/')[-1].split('_')
            _, thres_, tag_ = temp_split[0], temp_split[-2], temp_split[-1]
            sample_ = ''
            class_ = f.split('/')[-4]
            #tag_ = f.split('_')[-5]
            sv_type_ = ''
            model_ = model_n
            # print(f.split('_'))
            # if len(f.split('_')[-3].split('.')) < 3:
            #     continue
            # else:
            # thres_ = float(f.split('_')[-3].split('.')[1]) / 10
            batch_ = f.split('/')[-3]
            # info_list = [sample_, class_, tag_, sv_type_, thres_, batch_]
            
            # sub_f = os.path.join(root_folder, f)
            j_file = open(os.path.join(f, 'summary.txt'))
            data = json.loads(j_file.read())
            # print(data)
            prec = data['precision']
            recall = data['recall']
            f1 = data['f1']
            tp = data['TP-base']
            fp = data['FP']
            fn = data['FN']
            info_list = [tp, fp, fn, prec, recall, f1, sv_type_, batch_, tag_, sample_, class_, model_, float(thres_)]
            # print(info_list)
            # exit(-1)
            # print(thres_dict)
            # print()
            # print(float(thres_))
            # print(thres_dict.keys())
            if float(thres_) not in thres_dict.keys():
                # print("######################")
                thres_dict[float(thres_)] = []
                thres_dict[float(thres_)].append(info_list)
            else:
                # print("??????????????????????")
                thres_dict[float(thres_)].append(info_list)
            
            #"""
        # print(thres_dict)
        keys = list(thres_dict.keys())
        keys.sort()
        #print(keys)
        # exit(-1)
        i = 0
        for e in thres_dict:
            workbook = load_workbook(filename=args.output_excel)
            print(k)
            print(sheets_list[k])
            sheet = workbook[sheets_list[k]]
            # if class_ == '10xweb':
            #     starting_col_index = 1
            # else:
            #     starting_col_index = 10
            #"""
            
            # for i in range(len(keys)):
            # print()
            # print(i)
            sheet.cell(row=i+2, column=1, value = keys[i])
            # sheet.cell(row=i+2, column=13, value = keys[i])
            # print("?")
            # print(keys[i])
            # print(thres_dict[keys[i]])
            # print(thres_dict[keys[i]][0])
            # print(thres_dict[keys[i]][0][-2])
            for col in range(12):
                # print(thres_dict[keys[i]][0][col])
                sheet.cell(row=i+2, column=col+2, value = thres_dict[keys[i]][0][col])
            # if thres_dict[keys[i]][0][-2] == 'stlfr':
            #     for col in range(11):
            #         print(thres_dict[keys[i]][0][col])
            #         sheet.cell(row=i+2, column=col+14, value = thres_dict[keys[i]][0][col])
            #     for col in range(11):
            #         print(thres_dict[keys[i]][1][col])
            #         sheet.cell(row=i+2, column=col+2, value = thres_dict[keys[i]][1][col])
            # else:
            #     for col in range(11):
            #         print(thres_dict[keys[i]][1][col])
            #         sheet.cell(row=i+2, column=col+14, value = thres_dict[keys[i]][1][col])
            #     for col in range(11):
            #         print(thres_dict[keys[i]][0][col])
            #         sheet.cell(row=i+2, column=col+2, value = thres_dict[keys[i]][0][col])
            workbook.save(filename=args.output_excel)
            i += 1
        k += 1
        print("done")
